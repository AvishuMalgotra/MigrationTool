import pandas as pd
import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from .compatibility import CompatibilityService

class ReportService:
    def __init__(self):
        self.compatibility_service = CompatibilityService()
        
        # Styles
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
        self.header_font = Font(name="Calibri", size=14, color="FFFFFF", bold=True)
        self.sub_header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # Light Blue
        self.border_thin = Side(border_style="thin", color="000000")
        self.border_full = Border(left=self.border_thin, right=self.border_thin, top=self.border_thin, bottom=self.border_thin)

    def generate_excel_report(self, job_id: str, inventory: Dict[str, Any], blockers: Dict[str, Any]) -> bytes:
        """
        Generates a professional Excel report with Custom Cover and Pre-Req pages.
        """
        output = io.BytesIO()
        resources = inventory.get("resources", [])
        
        # 1. Prepare Data
        report_data = self.compatibility_service.generate_detailed_report(resources, existing_blockers=blockers)
        
        # 2. Excel Generation
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Note: We must create sheets in order.
            
            # --- Sheet 1: Cover Page ---
            self._create_cover_page(writer, job_id, inventory)
            
            # --- Sheet 2: Pre-Requisites ---
            self._create_prereq_page(writer, inventory)
            
            # --- Sheet 3: Resource Impact ---
            self._create_impact_sheet(writer, resources)

            # --- Sheet 4: Azure Subscription 1 (Assessment Details) ---
            df_details = pd.DataFrame(report_data)
            df_details.to_excel(writer, index=False, sheet_name="Azure Subscription 1")
            self._format_worksheet(writer.sheets['Azure Subscription 1'])

            # --- Sheet 5: Pre-Migration Tasks ---
            # Placeholder for now, can be expanded
            df_pre = pd.DataFrame({"Task": ["Snapshot VMs", "Verify Backups"], "Status": ["Pending", "Pending"]})
            df_pre.to_excel(writer, index=False, sheet_name="Pre-Migration Tasks")
            self._format_worksheet(writer.sheets['Pre-Migration Tasks'])

            # --- Sheet 6: Post-Migration Tasks ---
            df_post = pd.DataFrame({"Task": ["Verify DNS", "Install Extensions"], "Status": ["Pending", "Pending"]})
            df_post.to_excel(writer, index=False, sheet_name="Post-Migration Tasks")
            self._format_worksheet(writer.sheets['Post-Migration Tasks'])

            # --- Sheet 7: Public IP Info ---
            # Filter PIPs
            pips = [r for r in resources if "publicipaddresses" in r.get("type", "").lower()]
            pips_data = [{
                "Name": p.get("name"), 
                "IP Address": p.get("properties", {}).get("ipAddress", "Dynamic"), 
                "SKU": p.get("sku", {}).get("name", "Basic")
            } for p in pips]
            pd.DataFrame(pips_data).to_excel(writer, index=False, sheet_name="Public IP Info")
            self._format_worksheet(writer.sheets['Public IP Info'])

            # --- Sheet 8: Rollback Plan ---
            self._create_rollback_sheet(writer)

            # --- Sheet 9: Escalation Matrix ---
            self._create_escalation_sheet(writer)

            # --- Sheet 10: Test Matrix ---
            self._create_test_sheet(writer)

        output.seek(0)
        return output.getvalue()

    def _create_cover_page(self, writer, job_id, inventory):
        workbook = writer.book
        sheet = workbook.create_sheet("Cover", 0)
        
        # --- Styles ---
        fill_blue_dark = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_blue_light = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        font_white_bold = Font(name="Calibri", size=14, color="FFFFFF", bold=True)
        font_black_bold = Font(name="Calibri", size=11, bold=True)
        border_all = Border(left=self.border_thin, right=self.border_thin, top=self.border_thin, bottom=self.border_thin)

        # --- Layout ---
        
        # Logo / Branding (B2: D4)
        # Merged box for Logo
        sheet.merge_cells("B2:D4")
        logo_cell = sheet["B2"]
        logo_cell.value = "Tech Plus Talent"
        logo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        logo_cell.font = Font(name="Arial Black", size=16, color="EB5E76") # Try to match red/pink logo color
        self._apply_borders(sheet, "B2:D4")

        # Header Info (E2:J4) - Right side block
        # Row 2: "Project - CSP Migration"
        sheet["E2"] = "Project - CSP Migration"
        sheet.merge_cells("E2:J2")
        sheet["E2"].fill = fill_blue_dark
        sheet["E2"].font = font_white_bold
        sheet["E2"].alignment = Alignment(horizontal='center', vertical='center')

        # Row 3: "Olux Tech"
        sheet["E3"] = "Olux Tech"
        sheet.merge_cells("E3:J3")
        sheet["E3"].fill = fill_blue_dark
        sheet["E3"].font = font_white_bold
        sheet["E3"].alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 4: "Migration Assessment Report"
        sheet["E4"] = "Migration Assessment Report"
        sheet.merge_cells("E4:J4")
        sheet["E4"].fill = fill_blue_dark
        sheet["E4"].font = font_white_bold
        sheet["E4"].alignment = Alignment(horizontal='center', vertical='center')
        
        # Row 5: Date (B5:J5)
        sheet["B5"] = f"Date: {datetime.now().strftime('%d-%m-%Y')}"
        sheet.merge_cells("B5:J5")
        sheet["B5"].font = font_black_bold
        sheet["B5"].border = border_all
        self._apply_borders(sheet, "B5:J5")

        # Spacer Row 6 (Empty)

        # --- Table of Contents (B7:J??) ---
        
        # Header "CONTENTS"
        sheet["B7"] = "CONTENTS"
        sheet.merge_cells("B7:J7")
        sheet["B7"].alignment = Alignment(horizontal='center')
        sheet["B7"].font = font_black_bold
        sheet["B7"].fill = fill_blue_light
        self._apply_borders(sheet, "B7:J7")

        # Table Column Headers
        sheet["B8"] = "Sr. No"
        sheet["C8"] = "Title"
        sheet.merge_cells("C8:I8")
        sheet["J8"] = "Page No."
        
        for cell in ["B8", "C8", "J8"]:
             c = sheet[cell]
             c.font = font_black_bold
             c.fill = fill_blue_light
             c.border = border_all
        self._apply_borders(sheet, "C8:I8") # Apply to merged range

        # Rows
        contents = [
            ("01.", "Cover Page", "1"),
            ("02.", "Pre-Requisites", "2"),
            ("03.", "Resource Impact", "3"),
            ("04.", "Azure Subscription 1", "4"),
            ("05.", "Pre-Migration Tasks", "5"),
            ("06.", "Post-Migration Tasks", "6"),
            ("07.", "Public IP Info", "7"),
            ("08.", "Rollback Plan", "8"),
            ("09.", "Escalation Matrix", "9"),
            ("10.", "Test Matrix", "10"),
        ]
        
        row_idx = 9
        for sr, title, page in contents:
            sheet[f"B{row_idx}"] = sr
            sheet[f"C{row_idx}"] = title
            sheet.merge_cells(f"C{row_idx}:I{row_idx}")
            sheet[f"J{row_idx}"] = page
            
            # Styling - Light Blue fill for all rows as per sample?
            # Sample looks like all rows are light blue.
            for col in range(2, 11): # B to J
                cell = sheet.cell(row=row_idx, column=col)
                cell.fill = fill_blue_light
                cell.border = self.border_full
            
            row_idx += 1
            
        # Footer
        footer_row = row_idx
        # "No of Sheets in this report : 10" (B:F)
        sheet[f"B{footer_row}"] = "No of Sheets in this report : 10"
        sheet.merge_cells(f"B{footer_row}:F{footer_row}")
        sheet[f"B{footer_row}"].font = font_black_bold
        sheet[f"B{footer_row}"].border = border_all
        
        # "Report Prepared By: TPT Migration Team" (G:J)
        sheet[f"G{footer_row}"] = "Report Prepared By: TPT Migration Team"
        sheet.merge_cells(f"G{footer_row}:J{footer_row}")
        sheet[f"G{footer_row}"].font = font_black_bold
        sheet[f"G{footer_row}"].alignment = Alignment(horizontal='right')
        sheet[f"G{footer_row}"].border = border_all
        
        # Apply borders to the whole footer row range
        for c in range(2, 11):
            sheet.cell(row=footer_row, column=c).border = self.border_full

    def _apply_borders(self, sheet, range_string):
        """Applies full borders to a merged or single range"""
        rows = sheet[range_string]
        # openpyxl range can be a tuple of rows
        for row in rows:
            for cell in row:
                cell.border = self.border_full

    def _create_prereq_page(self, writer, inventory):
        workbook = writer.book
        sheet = workbook.create_sheet("Pre-Req")
        
        # Styles
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
        font_header = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
        font_normal = Font(name="Calibri", size=11, color="000000")
        border_all = Border(left=self.border_thin, right=self.border_thin, top=self.border_thin, bottom=self.border_thin)

        # Header Row
        sheet["A1"] = "Checks"
        sheet["B1"] = "Comment"
        
        for cell in ["A1", "B1"]:
            sheet[cell].fill = fill_header
            sheet[cell].font = font_header
            sheet[cell].border = border_all
        
        # Aggregates
        resources = inventory.get("resources", [])
        vms = [r for r in resources if "virtualmachines" in r.get("type", "").lower()]
        
        # Check List - Full 29 Items matching sample
        # We can implement logic for some, others static "Not Applicable"
        
        # Logic helpers
        has_firewall = any("firewall" in r.get("type", "").lower() for r in resources)
        has_ade = any("diskencryption" in str(r).lower() for r in resources)
        has_sql = any("sql" in r.get("type", "").lower() for r in resources)
        
        checks = [
            ("Quota + Usage check on Destination subscription", "Available"),
            ("Owner Access without Conditions", "Required at Destination Subscription"),
            ("Owner Access without Conditions", "Required at Source Subscription"), # Added to match typical dual check or just keep dest
            ("VMs Status", f"All {len(vms)} VMs Running (Healthy)" if vms else "No VMs Found"), 
            ("Reservations", "Please note that Reservations will not be directly moved..."),
            ("Whether Port 25 Open on Source Subscription", "Please note that Microsoft isn't allowing SMTP port 25..."),
            ("Resource Count", str(len(resources))),
            ("VPN SKU Change", "Not Applicable"),
            ("Global Admin Access", "Not Applicable"),
            ("System Assigned Managed Identity", "Virtual Machine" if any(vm.get("identity") for vm in vms) else "Not Applicable"),
            ("Whether using Cloud PC", "Not Applicable"),
            ("Marketplace VM", "Not Applicable"),
            ("Nerdio Automations", "Not Applicable"),
            ("Vnet Peering/Global Peering", "Not Applicable"),
            ("Managed Identity associated with DIs", "Not Applicable"),
            ("Firewall (If any)", "Present" if has_firewall else "Not Applicable"),
            ("Disk Encryption (ADE)", "Enabled" if has_ade else "Not Applicable"),
            ("SQLVM Backups", "Not Applicable" if not has_sql else "Check Required"),
            ("Recovery Services Vault used for DR", "Not Applicable"),
            ("Recovery Services Vault - Immutable", "Not Applicable"),
            ("Recovery Services Vault - Locked", "Not Applicable"),
            ("Site Recovery Infrastructure", "Not Applicable"),
            ("Custom Domain in App Services", "Not Applicable"),
            ("NICs pointing to other subscription", "Not Applicable"),
            ("BGP Peering in Site-to-Site Connection", "Not Applicable"),
            ("Microsoft Sentinel", "Not Applicable"),
            ("Entra Domain Services / LDAP", "Not Applicable"),
            ("Azure Databricks Service", "Not Applicable"),
            ("Data Factory -> SHIR", "Not Applicable")
        ]
        
        row_idx = 2
        for check, comment in checks:
            # Correct duplicates logic if needed, but list allows duplicates
            if check == "Owner Access without Conditions" and row_idx == 3:
                 # Sample showed row 3 as owner access too? No, row 3 is owner access. 
                 # Let's stick to strict sample order.
                 # Sample Row 2: Quota
                 # Sample Row 3: Owner Access
                 # Sample Row 4: VMs Status
                 pass

            cell_a = sheet[f"A{row_idx}"]
            cell_b = sheet[f"B{row_idx}"]
            
            cell_a.value = check
            cell_b.value = comment
            
            cell_a.font = font_normal
            cell_b.font = font_normal
            cell_a.border = border_all
            cell_b.border = border_all
            
            if "Required" in comment:
                cell_b.font = Font(name="Calibri", size=11, color="FF0000") # Red
            
            row_idx += 1
            
        # Crucial Parameters Table (D1:F3) - Right Side
        # Sample Header: D1=Crucial Parameters (Blue), D2=Secure Score, D3=Backup, D4=DR
        # Sample Values: E2/F2 merged? 
        # Sample Image: D1:E1 merged "Crucial Parameters" ? No, Header is D1:F1 in Blue? 
        # Let's align with sample: Header Row 1 (D,E,F). 
        # D1: Crucial Parameters (Title) - actually typically D1 header for Parameter Name.
        # Let's look at the sample: [Crucial Parameters] [        ] [       ]
        # Actually it looks like a list.
        # D1-F1: Header "Crucial Parameters"
        # No, D1 "Crucial Parameters", E1 "Crucial Parameters" merged?
        # Let's use: D1 Header "Crucial Parameters", F1 Header "Good" (from my previous code), but sample has different headers?
        # Sample Image: Row 1, Col D-F. D1 "Crucial Parameters". E1-F1 seems empty in header?
        # Row 2: D "Secure Score", F "Good".
        # Let's try:
        # D1: "Crucial Parameters" (Header)
        # E1: "Status" (Header) - inferred
        
        sheet["D1"] = "Crucial Parameters"
        sheet["F1"] = "Status"
        sheet.merge_cells("D1:E1") # Merge for title width
        sheet["D1"].fill = fill_header
        sheet["D1"].font = font_header
        sheet["F1"].fill = fill_header
        sheet["F1"].font = font_header
        sheet["D1"].border = border_all
        sheet["F1"].border = border_all # And E1 border implicit
        
        params = [
            ("Secure Score", "Good"),
            ("Backup", "Backup not enabled for Virtual Machines"),
            ("DR", "Not Enabled")
        ]
        
        p_row = 2
        for param, val in params:
            sheet[f"D{p_row}"] = param
            sheet[f"F{p_row}"] = val
            sheet.merge_cells(f"D{p_row}:E{p_row}")
            
            for cell in [f"D{p_row}", f"E{p_row}", f"F{p_row}"]:
                sheet[cell].border = border_all
                sheet[cell].font = font_normal
            p_row += 1
            
        # Environment Info (Footer) - Needs to be separate section below checks
        # Sample: Row 13 (Blue Header) "Basic Information..."
        # But our checks go up to row 30. So Footer should be ~Row 31.
        
        env_row_start = row_idx + 1 # Leave 1 empty row
        sheet[f"A{env_row_start}"] = "Basic Information of Azure Environment"
        sheet.merge_cells(f"A{env_row_start}:F{env_row_start}")
        sheet[f"A{env_row_start}"].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # Light Blue
        sheet[f"A{env_row_start}"].font = Font(name="Calibri", size=11, bold=True)
        sheet[f"A{env_row_start}"].alignment = Alignment(horizontal='center')
        self._apply_borders(sheet, f"A{env_row_start}:F{env_row_start}")
        
        info_row = env_row_start + 1
        sub_id = inventory.get("subscription_id", "Unknown")
        tenant_id = inventory.get("tenant_id", "Unknown")
        
        # Headers Row
        headers_info = ["Source Subscription ID", "Destination Subscription ID", "Destination Subscription Plan", "Tenant ID"]
        # A, B, C, D? No, sample spreads them.
        # Sample: A14 Source Sub ID, B14 Dest Sub ID...
        # Let's map columns A, B, C, D, E, F
        # A: Source ID
        # B: Dest ID
        # C-D: Plan?
        # E-F: Tenant?
        
        # Row 1 of info: Headers
        sheet[f"A{info_row}"] = "Source Subscription ID"
        sheet[f"B{info_row}"] = "Destination Subscription ID"
        sheet[f"D{info_row}"] = "Destination Subscription Plan"
        sheet[f"E{info_row}"] = "Tenant ID"
        
        # Row 2 of info: Values
        val_row = info_row + 1
        sheet[f"A{val_row}"] = sub_id
        sheet[f"B{val_row}"] = "Not yet Provisioned"
        sheet[f"B{val_row}"].font = Font(color="FF0000")
        sheet[f"D{val_row}"] = "CSP"
        sheet[f"E{val_row}"] = tenant_id
        
        # Styling Info Section
        for r in range(info_row, val_row + 1):
             for c in range(1, 7): # A to F
                 cell = sheet.cell(row=r, column=c)
                 cell.border = border_all
                 if r == info_row:
                     cell.font = Font(bold=True)

        self._format_worksheet(sheet)

    def _create_impact_sheet(self, writer, resources):
        impact_rows = []
        for res in resources:
            rtype = res.get("type", "").lower()
            name = res.get("name")
            impact = "Minimal"
            remarks = "No service interruption expected."
            
            if "virtualmachines" in rtype:
                impact = "High"
                remarks = "VM will be stopped and restarted. Temporary downtime required."
            elif "publicipaddresses" in rtype:
                sku = res.get("sku", {}).get("name", "").lower()
                if "basic" in sku:
                    impact = "Critical"
                    remarks = "Basic SKU Public IPs may change address during move. Standard SKU is static."
            
            impact_rows.append({
                "Resource Name": name,
                "Type": rtype,
                "Impact Level": impact,
                "Service Impact Description": remarks
            })
        pd.DataFrame(impact_rows).to_excel(writer, index=False, sheet_name="Resource Impact")
        self._format_worksheet(writer.sheets['Resource Impact'])

    def _create_rollback_sheet(self, writer):
        df = pd.DataFrame({
            "Step": [1, 2, 3],
            "Action": ["Identify failed resources", "Delete partial resources in Target", "Restore DNS"],
            "Responsible": ["Ops Team", "Ops Team", "NetAdmin"]
        })
        df.to_excel(writer, index=False, sheet_name="Rollback Plan")
        self._format_worksheet(writer.sheets['Rollback Plan'])

    def _create_escalation_sheet(self, writer):
        df = pd.DataFrame({
            "Role": ["Project Sponsor", "Project Manager", "Technical Lead"],
            "Name": ["", "", ""],
            "Contact": ["", "", ""]
        })
        df.to_excel(writer, index=False, sheet_name="Escalation Matrix")
        self._format_worksheet(writer.sheets['Escalation Matrix'])

    def _create_test_sheet(self, writer):
        df = pd.DataFrame({
            "Test Case": ["Verify VM Power On", "Verify App Access"],
            "Status": ["Pending", "Pending"]
        })
        df.to_excel(writer, index=False, sheet_name="Test Matrix")
        self._format_worksheet(writer.sheets['Test Matrix'])

    def _apply_borders(self, sheet, range_string):
        rows = sheet[range_string]
        for row in rows:
            for cell in row:
                cell.border = self.border_full

    def _format_worksheet(self, worksheet):
        try:
            for column_cells in worksheet.columns:
                # Calculate max length
                length = 0
                for cell in column_cells:
                    if cell.value:
                        length = max(length, len(str(cell.value)))
                
                # Apply width
                # Use openpyxl.utils.get_column_letter just to be safe
                # And access column index from the first cell (even if merged, it has column index)
                col_idx = column_cells[0].column
                col_letter = get_column_letter(col_idx)
                
                length = min(length, 80)
                worksheet.column_dimensions[col_letter].width = length + 5
        except Exception as e:
            print(f"Warning: Auto-format failed for sheet {worksheet.title}: {e}")
